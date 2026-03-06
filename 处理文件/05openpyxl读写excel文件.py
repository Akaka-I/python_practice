import datetime
import random

import openpyxl
wb=openpyxl.load_workbook('F:\python_demo\处理文件\\2022年股票数据_1.xlsx')
print(wb.sheetnames) #获取所有sheet的名字
sheet = wb.worksheets[0]
# 获得单元格的范围
print(sheet.dimensions)
print(sheet.max_row,sheet.max_column)

# 获取单元格的值
print(sheet.cell(2,3).value) #获取第二行第三列的值
print(sheet['A6'].value) #获取A6单元格的值
print(sheet['A2:C5']) #获取A2到C5的单元格范围

for row_ch in range(2,sheet.max_row+1):
    for col_ch in 'ABCDE':
        value=sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'),end=' \t')
        if type(value) == float:
            print(f'{value:.4f}',end=' \t')
        if type(value) == int:
            print(f'{value:<10d}',end=' \t')
    print()

""".....................写入excel操作.........................."""
wb=openpyxl.Workbook() #创建一个新的工作簿
sheet=wb.active #获取当前活跃的sheet
sheet.title='成绩表' #修改sheet的名字
titles=['姓名','语文','数学','英语']
for col_index,title in enumerate(titles):
    sheet.cell(1,col_index+1,title) #写入标题行
names=['张三','李四','王五']
for row_index,name in enumerate(names):
    sheet.cell(row_index+2,1,name) #写入姓名列
    for col_index in range(2,5):#写入成绩列
        sheet.cell(row_index+2,col_index,random.randint(60,101)) #写入成绩数据
wb.save('F:\python_demo\处理文件\\成绩表.xlsx') #保存工作簿
